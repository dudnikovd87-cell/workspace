#!/usr/bin/env python3
"""
Ежедневный отчёт Ozon Analytics
Запускается из cron в 8:00 МСК
"""
import json
import subprocess
import csv
import openpyxl
from openpyxl.styles import Font, Border, Side, PatternFill
from datetime import datetime
import os

CLIENT_ID = "1959914"
API_KEY = "3d9d2d17-a9f6-4fe5-8f3b-0aa12ad6acce"
REPORTS_DIR = "/root/.openclaw/workspace/reports"
ADS_CSV = f"{REPORTS_DIR}/ozon_ads_feb2026_raw.csv"

def get_analytics(date_from, date_to, max_attempts=3):
    """Выгрузить аналитику с несколькими попытками"""
    for attempt in range(1, max_attempts + 1):
        print(f"Попытка {attempt}/{max_attempts}...")
        cmd = [
            "curl", "-s", "-X", "POST", "https://api-seller.ozon.ru/v1/analytics/data",
            "-H", f"Client-Id: {CLIENT_ID}",
            "-H", f"Api-Key: {API_KEY}",
            "-H", "Content-Type: application/json",
            "-d", json.dumps({
                "date_from": date_from,
                "date_to": date_to,
                "metrics": ["revenue", "ordered_units", "hits_view", "hits_tocart", "session_view", "delivered_units", "returns", "cancellations"],
                "dimensions": ["sku"],
                "limit": 1000
            })
        ]
        result = subprocess.run(cmd, capture_output=True, text=True)
        
        if result.stdout and "result" in result.stdout:
            data = json.loads(result.stdout)
            if data.get('result', {}).get('data'):
                print("Выгрузка успешна!")
                return data
        
        print(f"Ошибка на попытке {attempt}")
    
    return None

def create_excel(analytics_data, output_path, ads_csv_path):
    # Читаем рекламные данные
    ads_data = {}
    with open(ads_csv_path, 'r') as f:
        for row in csv.DictReader(f):
            ads_data[row['ozon_sku']] = row
    
    print(f"Артикулов в рекламе: {len(ads_data)}")
    
    # Фильтруем - только те что в рекламе
    filtered_rows = []
    for item in analytics_data.get('result', {}).get('data', []):
        sku = item['dimensions'][0]['id']
        if sku in ads_data:  # Только если есть в рекламе
            m = item['metrics']
            ad = ads_data[sku]
            filtered_rows.append({
                'sku': sku,
                'article': ad.get('seller_article', ''),
                'revenue': m[0], 'ordered': m[1], 'hits_view': m[2], 'hits_tocart': m[3],
                'sessions': m[4], 'delivered': m[5], 'returns': m[6], 'cancellations': m[7],
                'ad_impr': float(ad.get('impr', 0)), 'ad_clicks': float(ad.get('clicks', 0)),
                'ad_ctr': float(ad.get('ctr_pct', 0)), 'ad_spend': float(ad.get('spend', 0)),
                'ad_orders': float(ad.get('orders', 0)), 'ad_sales': float(ad.get('sales', 0)),
                'ad_drr': float(ad.get('drr_pct', 0))
            })
    
    filtered_rows.sort(key=lambda x: x['revenue'], reverse=True)
    print(f"Отфильтровано: {len(filtered_rows)} товаров")
    
    # Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Озон Аналитика"
    
    headers = ["№","Артикул","Ozon SKU","Выручка","Заказано","Показы","В корзину","Сессии","Доставлено","Возвращено","Отменено","Выкуп(%)","Рекл.показы","Рекл.клики","Рекл.CTR(%)","Рекл.расход","Рекл.заказы","Рекл.продажи","ДРР(%)"]
    for col, h in enumerate(headers, 1):
        ws.cell(1, col, h)
    
    for i, r in enumerate(filtered_rows, 2):
        ws.cell(i, 1, i-1); ws.cell(i, 2, r['article']); ws.cell(i, 3, int(r['sku']))
        ws.cell(i, 4, r['revenue']); ws.cell(i, 5, r['ordered']); ws.cell(i, 6, r['hits_view'])
        ws.cell(i, 7, r['hits_tocart']); ws.cell(i, 8, r['sessions']); ws.cell(i, 9, r['delivered'])
        ws.cell(i, 10, r['returns']); ws.cell(i, 11, r['cancellations'])
        ws.cell(i, 12, round(r['delivered']/r['ordered']*100, 2) if r['ordered'] else 0)
        ws.cell(i, 13, r['ad_impr']); ws.cell(i, 14, r['ad_clicks']); ws.cell(i, 15, r['ad_ctr'])
        ws.cell(i, 16, r['ad_spend']); ws.cell(i, 17, r['ad_orders']); ws.cell(i, 18, r['ad_sales'])
        ws.cell(i, 19, r['ad_drr'])
    
    # Итого
    t = {c: sum(r[c] for r in filtered_rows) for c in ['revenue','ordered','hits_view','hits_tocart','sessions','delivered','returns','cancellations','ad_impr','ad_clicks','ad_spend','ad_orders','ad_sales']}
    i = len(filtered_rows) + 2
    ws.cell(i, 1, "ИТОГО")
    for col, key in [(4,'revenue'),(5,'ordered'),(6,'hits_view'),(7,'hits_tocart'),(8,'sessions'),(9,'delivered'),(10,'returns'),(11,'cancellations'),(13,'ad_impr'),(14,'ad_clicks'),(16,'ad_spend'),(17,'ad_orders'),(18,'ad_sales')]:
        ws.cell(i, col, t[key])
    ws.cell(i, 12, round(t['delivered']/t['ordered']*100, 2) if t['ordered'] else 0)
    ws.cell(i, 15, round(t['ad_clicks']/t['ad_impr']*100, 2) if t['ad_impr'] else 0)
    ws.cell(i, 19, round(t['ad_spend']/t['ad_sales']*100, 2) if t['ad_sales'] else 0)
    
    wb.save(output_path)
    return len(filtered_rows), t

def main():
    # Определяем период
    now = datetime.now()
    date_from = f"{now.year}-{now.month:02d}-01"
    date_to = now.strftime("%Y-%m-%d")
    
    # Для теста используем февраль
    date_from = "2026-02-01"
    date_to = "2026-02-28"
    
    print(f"Ozon Analytics: {date_from} - {date_to}")
    
    # Удаляем старые файлы
    raw_file = f"{REPORTS_DIR}/ozon_analytics_daily_raw.json"
    output_file = f"{REPORTS_DIR}/ozon_analytics_daily.xlsx"
    if os.path.exists(raw_file):
        os.remove(raw_file)
    if os.path.exists(output_file):
        os.remove(output_file)
    
    # Выгружаем (3 попытки)
    data = get_analytics(date_from, date_to)
    
    if not data:
        print("ОШИБКА: Не удалось выгрузить данные после 3 попыток")
        # Отправляю себе сообщение об ошибке - но это через меня
        exit(1)
    
    # Сохраняем raw
    with open(raw_file, 'w') as f:
        json.dump(data, f, indent=2)
    
    # Excel (только те что в рекламе)
    rows_count, totals = create_excel(data, output_file, ADS_CSV)
    
    print(f"Готово! Строк: {rows_count}")
    print(f"Выручка: {totals['revenue']:,}, Рекл.расход: {totals['ad_spend']:,.0f}")
    print(f"Файл: {output_file}")

if __name__ == "__main__":
    main()