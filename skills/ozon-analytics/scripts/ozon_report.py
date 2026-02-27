#!/usr/bin/env python3
"""
Скрипт для выгрузки данных Ozon и создания объединённого отчёта.
"""
import json
import csv
import os
import subprocess
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# Конфигурация
CLIENT_ID = "1959914"
API_KEY = "3d9d2d17-a9f6-4fe5-8f3b-0aa12ad6acce"
BASE_URL = "https://api-seller.ozon.ru"
REPORTS_DIR = "/root/.openclaw/workspace/reports"

def run_curl(cmd):
    """Выполнить curl и вернуть JSON"""
    result = subprocess.run(cmd, shell=True, capture_output=True, text=True)
    return json.loads(result.stdout)

def get_finance(date_from, date_to):
    """Выгрузить финансы"""
    cmd = f'''curl -s -X POST "{BASE_URL}/v1/finance/balance" \
      -H "Client-Id: {CLIENT_ID}" \
      -H "Api-Key: {API_KEY}" \
      -H "Content-Type: application/json" \
      -d '{{"date_from": "{date_from}", "date_to": "{date_to}"}}' '''
    return run_curl(cmd)

def get_analytics(date_from, date_to):
    """Выгрузить аналитику (Premium метрики)"""
    cmd = f'''curl -s -X POST "{BASE_URL}/v1/analytics/data" \
      -H "Client-Id: {CLIENT_ID}" \
      -H "Api-Key: {API_KEY}" \
      -H "Content-Type: application/json" \
      -d '{{"date_from": "{date_from}", "date_to": "{date_to}", \
      "metrics": ["revenue", "ordered_units", "hits_view", "hits_tocart", "session_view", "delivered_units", "returns", "cancellations"], \
      "dimensions": ["sku"], "limit": 1000}}' '''
    return run_curl(cmd)

def create_combined_report(analytics_data, ads_csv_path, output_path):
    """Создать объединённый отчёт"""
    # Читаем рекламные данные
    ads_data = {}
    with open(ads_csv_path, 'r') as f:
        reader = csv.DictReader(f)
        for row in reader:
            ads_data[row['ozon_sku']] = row

    # Обрабатываем аналитику
    rows = []
    for item in analytics_data['result']['data']:
        sku = item['dimensions'][0]['id']
        m = item['metrics']
        
        ad = ads_data.get(sku, {})
        
        rows.append({
            'sku': sku,
            'article': ad.get('seller_article', ''),
            'revenue': m[0],
            'ordered': m[1],
            'hits_view': m[2],
            'hits_tocart': m[3],
            'sessions': m[4],
            'delivered': m[5],
            'returns': m[6],
            'cancellations': m[7],
            'ad_impr': float(ad.get('impr', 0)),
            'ad_clicks': float(ad.get('clicks', 0)),
            'ad_ctr': float(ad.get('ctr_pct', 0)),
            'ad_spend': float(ad.get('spend', 0)),
            'ad_orders': float(ad.get('orders', 0)),
            'ad_sales': float(ad.get('sales', 0)),
            'ad_drr': float(ad.get('drr_pct', 0)),
        })

    # Сортируем по выручке
    rows.sort(key=lambda x: x['revenue'], reverse=True)

    # Создаём Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Озон Аналитика"

    headers = [
        "№", "Артикул", "Ozon SKU",
        "Выручка", "Заказано", "Показы", "В корзину", "Сессии", "Доставлено", "Возвращено", "Отменено", "Выкуп (%)",
        "Рекл. показы", "Рекл. клики", "Рекл. CTR (%)", "Рекл. расход", "Рекл. заказы", "Рекл. продажи", "ДРР (%)"
    ]

    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    for row_idx, r in enumerate(rows, 2):
        purchase_rate = round(r['delivered'] / r['ordered'] * 100, 2) if r['ordered'] else 0
        
        ws.cell(row=row_idx, column=1, value=row_idx-1)
        ws.cell(row=row_idx, column=2, value=r['article'])
        ws.cell(row=row_idx, column=3, value=int(r['sku']))
        ws.cell(row=row_idx, column=4, value=r['revenue'])
        ws.cell(row=row_idx, column=5, value=r['ordered'])
        ws.cell(row=row_idx, column=6, value=r['hits_view'])
        ws.cell(row=row_idx, column=7, value=r['hits_tocart'])
        ws.cell(row=row_idx, column=8, value=r['sessions'])
        ws.cell(row=row_idx, column=9, value=r['delivered'])
        ws.cell(row=row_idx, column=10, value=r['returns'])
        ws.cell(row=row_idx, column=11, value=r['cancellations'])
        ws.cell(row=row_idx, column=12, value=purchase_rate)
        ws.cell(row=row_idx, column=13, value=r['ad_impr'])
        ws.cell(row=row_idx, column=14, value=r['ad_clicks'])
        ws.cell(row=row_idx, column=15, value=r['ad_ctr'])
        ws.cell(row=row_idx, column=16, value=r['ad_spend'])
        ws.cell(row=row_idx, column=17, value=r['ad_orders'])
        ws.cell(row=row_idx, column=18, value=r['ad_sales'])
        ws.cell(row=row_idx, column=19, value=r['ad_drr'])

    # Форматирование
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for col in range(1, 20):
        ws.cell(row=1, column=col).fill = header_fill
        ws.cell(row=1, column=col).font = header_font
        ws.cell(row=1, column=col).alignment = Alignment(horizontal='center')
        for row in range(1, len(rows) + 2):
            ws.cell(row=row, column=col).border = thin_border

    wb.save(output_path)
    return len(rows)

if __name__ == "__main__":
    import sys
    if len(sys.argv) < 3:
        print("Usage: python ozon_report.py <date_from> <date_to>")
        sys.exit(1)
    
    date_from = sys.argv[1]
    date_to = sys.argv[2]
    
    print(f"Выгружаю данные за {date_from} - {date_to}...")
    
    # Выгружаем
    finance = get_finance(date_from, date_to)
    analytics = get_analytics(date_from, date_to)
    
    # Сохраняем
    with open(f"{REPORTS_DIR}/ozon_finance_{date_from}_{date_to}_raw.json", 'w') as f:
        json.dump(finance, f, indent=2)
    
    with open(f"{REPORTS_DIR}/ozon_analytics_{date_from}_{date_to}_raw.json", 'w') as f:
        json.dump(analytics, f, indent=2)
    
    print(f"Финансы и аналитика выгружены")
    print(f"Итого: Выручка={analytics['result']['totals'][0]}, Заказано={analytics['result']['totals'][1]}")
